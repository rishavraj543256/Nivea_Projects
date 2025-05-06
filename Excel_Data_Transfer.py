import pandas as pd
import shutil
from openpyxl import load_workbook
import os
from tkinter import Tk, filedialog, messagebox
from datetime import datetime
from openpyxl.styles import Border, Side, Alignment
import win32com.client
import pythoncom

# Update template file name to the correct one
TEMPLATE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "template", "template_delhivery.xlsx")

def select_file(title, filetypes=[("Excel files", "*.xlsx")]):
    """Open a dialog for user to select a file"""
    root = Tk()
    root.withdraw()  # Hide the main window
    file_path = filedialog.askopenfilename(
        title=title,
        filetypes=filetypes
    )
    return file_path

def copy_template_file(new_file_path):
    """Copy the template Excel file to a new location"""
    try:
        if not os.path.exists(TEMPLATE_FILE):
            # Try alternative paths
            alt_paths = [
                os.path.join(os.path.dirname(os.path.abspath(__file__)), "template_delhivery.xlsx"),
                os.path.join(os.path.dirname(os.path.abspath(__file__)), "template", "template_delhivery.xlsx"),
                "./template/template_delhivery.xlsx",
                "template/template_delhivery.xlsx"
            ]
            
            for path in alt_paths:
                if os.path.exists(path):
                    print(f"Using template file from alternate path: {path}")
                    shutil.copy2(path, new_file_path)
                    print(f"Template file copied successfully to: {new_file_path}")
                    return True
                    
            # If we got here, none of the paths worked
            raise FileNotFoundError(f"Template file not found at: {TEMPLATE_FILE} or any alternative locations")
            
        shutil.copy2(TEMPLATE_FILE, new_file_path)
        print(f"Template file copied successfully to: {new_file_path}")
        return True
    except Exception as e:
        print(f"Error copying template file: {e}")
        messagebox.showerror("Error", f"Error copying template file: {e}")
        return False

def get_data_from_source(source_file, sheet_name="Sheet2"):
    """Read data from the source Excel file"""
    try:
        df = pd.read_excel(source_file, sheet_name=sheet_name)
        return df
    except Exception as e:
        print(f"Error reading source file: {e}")
        return None
    
def remove_empty_rows_between_labels(sheet):
    try:
        # Find the 'Particulars' label row
        particulars_row = None
        total_row = None
        
        # Search for 'Particulars' label
        for row in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=2).value  # Column B
            if cell_value and "Particulars" in str(cell_value):
                particulars_row = row
                break
        
        if particulars_row is None:
            print("Could not find 'Particulars' label")
            return False
        
        # Search for 'Total' label starting from particulars_row
        for row in range(particulars_row, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=3).value  # Column C
            if cell_value and "Total" in str(cell_value):
                total_row = row
                break
        
        if total_row is None:
            print("Could not find 'Total' label")
            return False
        
        # Iterate through rows between particulars_row and total_row
        # and delete empty rows
        rows_to_delete = []
        for row in range(particulars_row + 1, total_row):
            is_empty = True
            for col in range(1, sheet.max_column + 1):  # Check all columns
                if sheet.cell(row=row, column=col).value is not None:
                    is_empty = False
                    break
            
            if is_empty:
                rows_to_delete.append(row)
        
        # Delete rows in reverse order to avoid shifting issues
        for row in sorted(rows_to_delete, reverse=True):
            sheet.delete_rows(row)
        
        return True
    
    except Exception as e:
        print(f"Error removing empty rows: {e}")
        return False

def update_particulars_data(new_sheet, invoice_data):
    """Update the Particulars, LR Number, and Amt (In INR) columns in the output sheet"""
    try:
        # Define the expected particulars in the output sheet in the correct order
        output_particulars_order = [
            "Rate Difference",
            "Damage",
            "Shortage",
            "Damage & Shortage"
        ]
        # Define border style
        thin_border = Border(left=Side(style='medium'), 
                         right=Side(style='medium'), 
                         top=Side(style='medium'), 
                         bottom=Side(style='medium'))
        
        # Find the starting row for particulars data (assuming it starts after headers)
        start_row = None
        for row in range(1, new_sheet.max_row + 1):
            if new_sheet.cell(row=row, column=1).value == "Sr":
                start_row = row + 3  # Skip the header rows
                break
        
        if start_row is None:
            print("Could not find starting row for particulars data")
            return False, []
        
        # Clear any existing data in the particulars section
        for row in range(start_row, start_row + 10):  # Clear next 10 rows to be safe
            for col in range(1, 6):  # Columns A to E
                new_sheet.cell(row=row, column=col).value = None
        
        # Process each particular type in the defined order
        row_counter = 0
        total_amount = 0
        rows_with_lr_data = []
        
        for particular in output_particulars_order:
            target_row = start_row + row_counter
            
            # Filter source data for this invoice and particular
            filtered_data = invoice_data[invoice_data['Particular'] == particular]
            
            # Only create a row if there's data for this particular
            if not filtered_data.empty:
                # Set Sr. No.
                new_sheet.cell(row=target_row, column=1).value = row_counter + 1
                
                # Set the particular in column B
                new_sheet.cell(row=target_row, column=2).value = particular
                
                # Get all LR Numbers for this particular
                lr_numbers = filtered_data['LR Number'].dropna().unique().tolist()
                # Join multiple LR numbers with comma if needed
                lr_numbers_str = ", ".join(str(num) for num in lr_numbers)
                new_sheet.cell(row=target_row, column=3).value = lr_numbers_str if lr_numbers_str else "NA"
                
                # Track rows that have LR data for later height adjustment
                if len(lr_numbers) > 1:
                    rows_with_lr_data.append(target_row)
                
                # Get the sum of amounts for this particular
                part_amount = filtered_data['Amt.'].sum()
                new_sheet.cell(row=target_row, column=4).value = part_amount
                new_sheet.cell(row=target_row, column=5).value = part_amount
                new_sheet.cell(row=target_row, column=5).border = thin_border
                total_amount += part_amount
                
                row_counter += 1

        # Add the total row at row 26
        total_row = 26
        new_sheet.cell(row=total_row, column=3).value = "Total"
        new_sheet.cell(row=total_row, column=5).value = total_amount
        
        # Add PAN number at the bottom
        new_sheet['A28'] = 'PAN Number   : AACCN1990P'
        
        return True, rows_with_lr_data
    except Exception as e:
        print(f"Error updating particulars data: {e}")
        return False, []

def create_sheet_for_invoice(wb, template_sheet, invoice_data, sheet_name, mappings):
    """
    Create a new sheet for a specific invoice and update its values
    
    Args:
        wb: Workbook object
        template_sheet: Template sheet to copy format from
        invoice_data: DataFrame containing data for this invoice
        sheet_name: Name for the new sheet
        mappings: List of mapping dictionaries
    """
    try:
        # Create a copy of the template sheet
        new_sheet = wb.copy_worksheet(template_sheet)
        new_sheet.title = sheet_name
        
        # Process each mapping
        for mapping in mappings:
            source_column = mapping.get('source_column')
            target_label = mapping['target_label']
            target_column = mapping['target_column']
            
            # Get the value based on mapping type
            if target_label == "Date":  # Special case for current date
                source_value = datetime.now().strftime("%d/%m/%Y")
            elif source_column and source_column in invoice_data.columns:
                source_value = invoice_data[source_column].iloc[0] if not invoice_data.empty else ""
            else:
                continue  # Skip if no source column and not a special case

            # Special case for "Ref" label - prepend "TCN-"
            if target_label == "Ref":
                source_value = f"TCN-{source_value}"

            # Find the row with the target label
            target_row = None
            for row in range(1, new_sheet.max_row + 1):
                for col in range(1, new_sheet.max_column + 1):
                    cell_value = new_sheet.cell(row=row, column=col).value
                    if cell_value and str(cell_value).strip() == str(target_label).strip():
                        target_row = row
                        break
                if target_row is not None:
                    break
            
            # Update the value if label was found
            if target_row is not None:
                target_cell = f"{target_column}{target_row}"
                new_sheet[target_cell] = source_value
                print(f"Updated {target_label} with value: {source_value} in sheet {sheet_name}")
            else:
                print(f"Warning: Label '{target_label}' not found in sheet {sheet_name}")
        
        # Update the particulars data
        update_success, rows_with_lr_data = update_particulars_data(new_sheet, invoice_data)
        if not update_success:
            print(f"Failed to update particulars data for sheet {sheet_name}")
        else:
            # Remove empty rows after data is populated
            remove_empty_rows_between_labels(new_sheet)
            
            # Find and format the LR Number header cell
            lr_header_cell = None
            for row in range(1, new_sheet.max_row + 1):
                for col in range(1, new_sheet.max_column + 1):
                    cell_value = new_sheet.cell(row=row, column=col).value
                    if cell_value and "LR Number" in str(cell_value):
                        lr_header_cell = (row, col)
                        break
                if lr_header_cell:
                    break
            
            # Set column width for LR Number column if found
            if lr_header_cell:
                lr_col = lr_header_cell[1]
                # Set column width wide enough for typical LR numbers
                new_sheet.column_dimensions[chr(64 + lr_col)].width = 25
            
            # Adjust row heights and set wrap text for LR Number cells
            for row in rows_with_lr_data:
                # Get the LR Number cell
                lr_cell = new_sheet.cell(row=row, column=3)
                
                # Set text wrapping
                lr_cell.alignment = Alignment(wrap_text=True, vertical='center')
                
                # Adjust row height based on content length
                content_length = len(str(lr_cell.value))
                if content_length > 30:
                    # Calculate appropriate height - adjust these values as needed
                    char_height = 2  # Approximate height per character
                    additional_height = min(80, (content_length // 20) * 15)  # Cap at 80 to avoid extreme heights
                    new_sheet.row_dimensions[row].height = 20 + additional_height
                else:
                    new_sheet.row_dimensions[row].height = 20
        
        return True
    except Exception as e:
        print(f"Error creating sheet for invoice: {e}")
        return False

def find_and_update_values(target_file, source_data, mappings):
    """
    Update the target Excel file with values from source data based on mappings
    Creates a new sheet for each unique invoice number
    """
    try:
        # Load the workbook
        wb = load_workbook(target_file)
        template_sheet = wb.active
        
        # Get unique invoice numbers
        invoice_numbers = source_data['Invoice Number'].unique()
        print(f"\nFound {len(invoice_numbers)} unique invoice numbers")
        
        # Create a sheet for each invoice number
        for i, invoice_num in enumerate(invoice_numbers, 1):
            # Filter data for this invoice
            invoice_data = source_data[source_data['Invoice Number'] == invoice_num]
            
            # Create sheet name (clean invoice number to make valid sheet name)
            sheet_name = f"Invoice_{str(invoice_num)}"
            # Remove invalid characters from sheet name
            sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in ['_', '-'])[:31]
            
            print(f"\nProcessing invoice {i} of {len(invoice_numbers)}: {invoice_num}")
            if create_sheet_for_invoice(wb, template_sheet, invoice_data, sheet_name, mappings):
                print(f"Successfully created sheet for invoice: {invoice_num}")
            else:
                print(f"Failed to create sheet for invoice: {invoice_num}")
        
        # Remove the template sheet
        wb.remove(template_sheet)
        
        # Save the workbook
        wb.save(target_file)
        print("\nTarget file updated successfully")
        return True
    except Exception as e:
        print(f"Error updating target file: {e}")
        return False

def convert_sheets_to_pdf(excel_file):
    """Convert each sheet in the Excel file to a separate PDF"""
    excel = None
    wb = None
    try:
        print("\nStarting PDF conversion...")
        # Initialize COM objects
        pythoncom.CoInitialize()
        
        # Create Excel application object with error handling
        try:
            excel = win32com.client.DispatchEx('Excel.Application')
        except Exception as e:
            print(f"Error creating Excel application: {e}")
            try:
                excel = win32com.client.gencache.EnsureDispatch('Excel.Application')
            except Exception as e2:
                print(f"Error with alternative Excel initialization: {e2}")
                raise Exception("Could not initialize Excel application")
        
        try:
            excel.Visible = False
            excel.DisplayAlerts = False
        except Exception as e:
            print(f"Warning: Could not set Excel visibility/alerts (this is okay): {e}")
        
        # Get the directory of the Excel file
        output_dir = os.path.dirname(excel_file)
        pdf_dir = os.path.join(output_dir, "PDFs")
        
        # Create PDFs directory if it doesn't exist
        if not os.path.exists(pdf_dir):
            os.makedirs(pdf_dir)
        
        # Open the workbook with absolute path
        abs_path = os.path.abspath(excel_file)
        print(f"Opening workbook: {abs_path}")
        try:
            wb = excel.Workbooks.Open(abs_path)
        except Exception as e:
            print(f"Error opening workbook: {e}")
            raise
        
        # Convert each sheet to PDF
        conversion_success = False
        for sheet in wb.Sheets:
            try:
                sheet.Select()
                
                # Additional preparation for the sheet
                print(f"Preparing sheet '{sheet.Name}' for PDF conversion...")
                
                # AutoFit rows (this actually does what Excel's "AutoFit Row Height" does)
                try:
                    # Special handling for LR Number column - ensure it's fully visible
                    # Try to find the LR Number column
                    lr_col_index = None
                    for col in range(1, sheet.UsedRange.Columns.Count + 1):
                        for row in range(1, 10):  # Check first few rows for headers
                            cell_value = sheet.Cells(row, col).Value
                            if cell_value and isinstance(cell_value, str) and "LR Number" in cell_value:
                                lr_col_index = col
                                break
                        if lr_col_index:
                            break
                    
                    # If LR Number column found, ensure it has proper width and wrap text
                    if lr_col_index:
                        print(f"Found LR Number column at index {lr_col_index}")
                        
                        # Set wider column width for LR Number column
                        sheet.Columns(lr_col_index).ColumnWidth = 30
                        
                        # Apply word wrap to the entire column 
                        sheet.Columns(lr_col_index).WrapText = True
                        
                        # Set page setup options before printing
                        sheet.PageSetup.PaperSize = 9  # xlPaperA4
                        sheet.PageSetup.Orientation = 2  # xlLandscape (if needed)
                        sheet.PageSetup.FitToPagesWide = 1  # Fit to 1 page wide
                        sheet.PageSetup.FitToPagesTall = False  # Let height adjust as needed
                        sheet.PageSetup.Zoom = False  # Disable zoom to enable fit to page
                        
                        # Auto-fit row heights - Apply to entire sheet
                        sheet.UsedRange.Rows.AutoFit()
                        
                        # Scan for particularly long content in LR Number column
                        for row in range(1, 30):  # Look at all possible data rows
                            cell_value = sheet.Cells(row, lr_col_index).Value
                            if cell_value and isinstance(cell_value, str) and "," in cell_value:
                                # This is likely a cell with multiple LR numbers - count how many
                                comma_count = cell_value.count(",")
                                # Set extra height based on number of commas (each roughly represents a line break)
                                min_height = 15 * (comma_count + 1)  # Base height per line
                                
                                # Compare with current height and increase if needed
                                current_height = sheet.Rows(row).RowHeight
                                if current_height < min_height:
                                    sheet.Rows(row).RowHeight = min_height
                                    print(f"Adjusted row {row} height to {min_height} for multiple LR numbers")
                
                except Exception as prep_error:
                    print(f"Warning: Row height adjustment failed, but continuing: {prep_error}")
                
                # Create PDF filename based on sheet name
                # Clean the sheet name to remove invalid characters
                clean_name = "".join(c for c in sheet.Name if c.isalnum() or c in ['-', '_', ' '])
                pdf_name = f"{clean_name}.pdf"
                pdf_path = os.path.join(pdf_dir, pdf_name)
                
                print(f"Converting sheet '{sheet.Name}' to PDF...")
                
                # Try different methods of PDF export
                try:
                    # Method 1: ExportAsFixedFormat with specific parameters for quality
                    sheet.ExportAsFixedFormat(
                        Type=0,  # xlTypePDF
                        Filename=pdf_path,
                        Quality=0,  # xlQualityStandard
                        IncludeDocProperties=True,
                        IgnorePrintAreas=False,
                        OpenAfterPublish=False
                    )
                except Exception as e1:
                    print(f"First export method failed: {e1}")
                    try:
                        # Method 2: Alternative export method
                        wb.ActiveSheet.ExportAsFixedFormat(0, pdf_path)
                    except Exception as e2:
                        print(f"Second export method failed: {e2}")
                        raise
                
                print(f"Created PDF: {pdf_name}")
                conversion_success = True
                
            except Exception as sheet_error:
                print(f"Error converting sheet '{sheet.Name}' to PDF: {sheet_error}")
                continue
        
        if not conversion_success:
            raise Exception("No sheets were successfully converted to PDF")
        
        print("\nPDF conversion completed successfully!")
        print(f"PDFs are saved in: {pdf_dir}")
        return True
        
    except Exception as e:
        print(f"Error converting to PDF: {e}")
        return False
        
    finally:
        # Cleanup in finally block to ensure it always runs
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
            if excel is not None:
                excel.Quit()
            pythoncom.CoUninitialize()
        except:
            pass

def main():
    try:
        # Select the source data file
        print("\nPlease select the source data Excel file...")
        source_file = select_file("Select Source Data Excel File")
        if not source_file:
            print("No source file selected. Exiting...")
            return

        # Create new file name based on template
        template_dir = os.path.dirname(TEMPLATE_FILE)
        template_name = os.path.basename(TEMPLATE_FILE)
        
        # Create the template directory if it doesn't exist
        output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            
        new_file_name = f"Filled_{template_name}"
        new_file_path = os.path.join(output_dir, new_file_name)

        # Copy template to new file
        if not copy_template_file(new_file_path):
            return

        # Read source data
        source_data = get_data_from_source(source_file)
        if source_data is None:
            return

        # Define mappings (can be expanded or loaded from a configuration file)
        mappings = [
            {
                'source_column': 'Invoice Number',
                'target_label': 'Bill no',
                'target_column': 'E'
            },
            {
                'source_column': 'Invoice Number',
                'target_label': 'Ref',
                'target_column': 'E'
            },
            {
                'source_column': 'Transporter Name',
                'target_label': 'KIND  ATTN : ',
                'target_column': 'B'
            },
            {
                'source_column': 'Invoice Date',
                'target_label': 'BILL DT',
                'target_column': 'E'
            },
            {
                'target_label': 'Date',
                'target_column': 'E'
            }
        ]
        
        # Update the new file with values
        if find_and_update_values(new_file_path, source_data, mappings):
            print(f"\nExcel file created successfully at: {new_file_path}")
            
            # Convert all sheets to PDF
            print("\nConverting sheets to PDF...")
            if convert_sheets_to_pdf(new_file_path):
                print("All sheets have been converted to PDF successfully!")
            else:
                print("Failed to convert sheets to PDF!")
                
            return new_file_path
        else:
            print("\nProcess failed!")
            return None

    except Exception as e:
        print(f"An error occurred: {e}")
        return None

if __name__ == "__main__":
    main()
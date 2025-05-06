import os
import pandas as pd
from tkinter import Tk, filedialog, messagebox
import glob
import datetime
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, PatternFill

def select_folder():
    """Open a dialog for user to select a folder"""
    root = Tk()
    root.withdraw()  # Hide the main window
    folder_path = filedialog.askdirectory(title="Select Folder Containing CSV Files")
    return folder_path

def select_existing_file():
    """Open a dialog for user to select an existing consolidated file"""
    root = Tk()
    root.withdraw()  # Hide the main window
    file_path = filedialog.askopenfilename(
        title="Select Existing Consolidated File",
        filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")]
    )
    return file_path

def ask_user_preference():
    """Ask user if they want to create new file or append to existing"""
    root = Tk()
    root.withdraw()
    response = messagebox.askquestion(
        "Output Preference",
        "Do you want to append to an existing consolidated file?\n\n" +
        "Click 'Yes' to append to existing file\n" +
        "Click 'No' to create a new file"
    )
    return response == 'yes'

def find_combined_files(folder_path):
    """Find all CSV files starting with 'combined_' recursively in all subfolders"""
    files = []
    
    # Walk through all subdirectories
    for root, dirs, filenames in os.walk(folder_path):
        #import ipdb;ipdb.set_trace()
        if (root.split('\\')[-1])=='delhivery':
            for filename in filenames:
                #import ipdb;ipdb.set_trace()
                
                # Check if file starts with 'combined_' and has csv extension
                if not filename.startswith('combined_') and filename.endswith('.csv'):
                    full_path = os.path.join(root, filename)
                    #import ipdb;ipdb.set_trace()
                    files.append(full_path)
                    print(f"Found file: {full_path}")
    
    return files

def create_consolidated_folder(base_folder):
    """Create a consolidated folder if it doesn't exist"""
    consolidated_folder = os.path.join(base_folder, "consolidated")
    if not os.path.exists(consolidated_folder):
        os.makedirs(consolidated_folder)
    return consolidated_folder

def create_sheet2_template():
    """Create an empty DataFrame for Sheet2 with the required columns"""
    columns = [
        'Invoice Number', 'Invoice Date', 'Transporter Name', 'LR Number',
        'Nart No.', 'Per piece Value', 'Pack Size', 'Case', 'Loose Pieces',
        'Total Pieces', 'Amt.', 'Particular'
    ]
    return pd.DataFrame(columns=columns)

def combine_csv_files(files, existing_df=None):
    """Combine multiple CSV files into a single DataFrame"""
    if not files:
        raise ValueError("No CSV files found starting with 'combined_'")
    
    # List to store all dataframes
    dfs = []
    
    # Required columns
    required_columns = [
        'Invoice Number(s)', 'Pickup Date', 'Delivered Date', 'Total Amount',
        'Consignee Name', 'LRN', 'Origin City', 'Destination City'
    ]
    
    # Read each CSV file
    for file in files:
        try:
            print(f"Reading file: {os.path.basename(file)}")
            try:
                df = pd.read_csv(file, encoding='utf-8')
            except UnicodeDecodeError:
                df = pd.read_csv(file, encoding='latin1')
            
            # Keep only required columns if they exist
            available_columns = [col for col in required_columns if col in df.columns]
            if not available_columns:
                print(f"Warning: No required columns found in {file}")
                continue
                
            df = df[available_columns]
            
            # Add source information
            df['Source_File'] = os.path.basename(file)
            df['Source_Folder'] = os.path.dirname(file)
            
            # Add Transporter Name column
            df['Transporter Name'] = 'Delhivery Limited'
            
            # Convert Pickup Date to datetime and extract month
            if 'Pickup Date' in df.columns:
                df['Pickup Date'] = pd.to_datetime(df['Pickup Date'], errors='coerce')
                df['Invoice Month'] = df['Pickup Date'].dt.strftime('%B %Y')
            
            # Add new required columns with empty values
            df['Vouched By'] = ''
            df['LR Vouching Date'] = ''
            df['Invoice Vouching Date'] = ''
            df['LR Query'] = ''
            df['Invoice Query'] = ''
            df['LR Loss Value'] = ''
            df['Freight Status'] =''
            df['LR Status'] =''
            df['Invoice Status'] =''
            
            dfs.append(df)
        except Exception as e:
            print(f"Error reading {file}: {e}")
    
    if not dfs:
        raise ValueError("No data could be read from the CSV files")
    
    # Combine all new dataframes
    new_data_df = pd.concat(dfs, ignore_index=True)
    
    if existing_df is not None:
        # Create a unique identifier for each row by combining Source_File and Source_Folder
        existing_df['file_folder_key'] = existing_df['Source_File'] + '|' + existing_df['Source_Folder']
        new_data_df['file_folder_key'] = new_data_df['Source_File'] + '|' + new_data_df['Source_Folder']
        
        # Filter out duplicates from new data
        new_unique_data = new_data_df[~new_data_df['file_folder_key'].isin(existing_df['file_folder_key'])]
        
        # Drop the temporary key column
        new_unique_data = new_unique_data.drop('file_folder_key', axis=1)
        existing_df = existing_df.drop('file_folder_key', axis=1)
        
        # Combine existing and new unique data
        print(f"Found {len(new_unique_data)} new unique entries to append")
        combined_df = pd.concat([existing_df, new_unique_data], ignore_index=True)
        
        # Check for duplicate LRs in the combined dataset
        combined_df['Duplicate LR'] = combined_df.duplicated(subset=['LRN'], keep=False).map({True: 'Yes', False: 'No'})
        
        return combined_df
    
    # For new consolidated file, check for duplicate LRs
    new_data_df['Duplicate LR'] = new_data_df.duplicated(subset=['LRN'], keep=False).map({True: 'Yes', False: 'No'})
    
    return new_data_df

def fill_invoice_dates(consolidated_file):
    """
    Fill the Invoice Date column in Sheet2 based on matching Invoice Numbers 
    from a user-selected Excel file.
    
    Args:
        consolidated_file: Path to the consolidated Excel file with Sheet2
        
    Returns:
        int: Count of updated dates, 0 if no updates were made
    """
    try:
        # Ask user if they want to fill Invoice Dates now
        root = Tk()
        root.withdraw()
        response = messagebox.askquestion(
            "Fill Invoice Dates",
            "Do you want to fill the Invoice Date column now?\n\n" +
            "Note: This requires Invoice Numbers to be filled in Sheet2 manually first.\n" +
            "Click 'Yes' to proceed or 'No' to skip."
        )
        
        if response != 'yes':
            print("Skipping Invoice Date automation.")
            return 0
            
        # Check if Sheet2 has data
        try:
            sheet2_df = pd.read_excel(consolidated_file, sheet_name='Sheet2')
            if sheet2_df.empty or all(sheet2_df['Invoice Number'].isna()):
                messagebox.showwarning(
                    "Empty Data", 
                    "Sheet2 is empty or no Invoice Numbers are filled yet. Please fill Invoice Numbers first."
                )
                return 0
        except Exception as e:
            print(f"Error reading Sheet2: {e}")
            messagebox.showerror("Error", f"Could not read Sheet2: {e}")
            return 0
            
        # Ask user to select the source file for Invoice Dates
        root = Tk()
        root.withdraw()
        source_file = filedialog.askopenfilename(
            title="Select Excel file containing Invoice Dates",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")]
        )
        
        if not source_file:
            print("No source file selected. Skipping Invoice Date automation.")
            return 0
            
        # Read the source file
        try:
            if source_file.endswith('.xlsx'):
                # Try each sheet until we find one with the Invoice Number and Invoice Date columns
                source_df = None
                xl = pd.ExcelFile(source_file)
                sheet_names = xl.sheet_names
                
                for sheet in sheet_names:
                    temp_df = pd.read_excel(source_file, sheet_name=sheet)
                    # Check if required columns exist (try different potential column names)
                    invoice_num_cols = [col for col in temp_df.columns if 'invoice number' in col.lower() or 'invoice no' in col.lower()]
                    invoice_date_cols = [col for col in temp_df.columns if 'invoice date' in col.lower() or 'date' in col.lower()]
                    
                    if invoice_num_cols and invoice_date_cols:
                        source_df = temp_df
                        invoice_num_col = invoice_num_cols[0]
                        invoice_date_col = invoice_date_cols[0]
                        print(f"Found Invoice Number column: {invoice_num_col} and Invoice Date column: {invoice_date_col} in sheet: {sheet}")
                        break
                
                if source_df is None:
                    messagebox.showerror("Error", "Could not find sheets with both Invoice Number and Invoice Date columns.")
                    return 0
            else:  # CSV file
                try:
                    # Try multiple encodings to handle potential encoding issues
                    encodings_to_try = ['utf-8', 'latin1', 'cp1252', 'ISO-8859-1']
                    
                    for encoding in encodings_to_try:
                        try:
                            print(f"Trying to read CSV with {encoding} encoding...")
                            source_df = pd.read_csv(source_file, encoding=encoding)
                            break  # If successful, break the loop
                        except UnicodeDecodeError:
                            if encoding == encodings_to_try[-1]:  # Last encoding to try
                                raise  # Re-raise the exception if all encodings fail
                            continue  # Try the next encoding
                        except Exception as e:
                            raise Exception(f"Error reading CSV: {str(e)}")
                    
                    # Check if required columns exist
                    invoice_num_cols = [col for col in source_df.columns if 'invoice number' in col.lower() or 'invoice no' in col.lower()]
                    invoice_date_cols = [col for col in source_df.columns if 'invoice date' in col.lower() or 'date' in col.lower()]
                    
                    if not invoice_num_cols or not invoice_date_cols:
                        # If we can't find columns based on their names, try to use the first and second columns
                        if len(source_df.columns) >= 2:
                            print("Using first column as Invoice Number and second as Invoice Date.")
                            invoice_num_col = source_df.columns[0]
                            invoice_date_col = source_df.columns[1]
                        else:
                            messagebox.showerror("Error", "Source file does not have enough columns to identify Invoice Number and Invoice Date.")
                            return 0
                    else:
                        invoice_num_col = invoice_num_cols[0]
                        invoice_date_col = invoice_date_cols[0]
                        
                    print(f"Using columns: '{invoice_num_col}' for Invoice Number and '{invoice_date_col}' for Invoice Date")
                    
                except Exception as e:
                    print(f"Error reading CSV file: {e}")
                    messagebox.showerror("Error", f"Could not read source file: {e}")
                    return 0
            
            # Create mapping of invoice numbers to dates
            invoice_date_map = {}
            for _, row in source_df.iterrows():
                try:
                    # Handle different data types for invoice numbers
                    invoice_num = row[invoice_num_col]
                    
                    # Skip empty values
                    if pd.isna(invoice_num):
                        continue
                    
                    # Convert numeric invoice numbers to string properly
                    if isinstance(invoice_num, (int, float)):
                        invoice_num = str(int(invoice_num))
                    else:
                        invoice_num = str(invoice_num).strip()
                    
                    if not invoice_num:  # Skip empty values after stripping
                        continue
                    
                    # Process date value if it exists
                    date_value = row[invoice_date_col]
                    if pd.isna(date_value):
                        continue
                    
                    # Format the date value to DD-MM-YYYY
                    try:
                        # Convert to datetime first, then format
                        parsed_date = pd.to_datetime(date_value)
                        formatted_date = parsed_date.strftime('%d-%m-%Y')
                        invoice_date_map[invoice_num] = formatted_date
                    except Exception as e:
                        print(f"Warning: Could not format date for invoice {invoice_num}: {e}")
                        # Use original value as fallback
                        invoice_date_map[invoice_num] = str(date_value)
                    
                except Exception as e:
                    print(f"Error processing row: {e}")
                    continue
            
            if not invoice_date_map:
                messagebox.showwarning("Warning", "No valid invoice number and date pairs found in source file.")
                return 0
            
            # Display summary of what we found
            print(f"Found {len(invoice_date_map)} invoice numbers with dates in the source file")
            print("First 5 entries in the mapping:")
            for i, (invoice, date) in enumerate(list(invoice_date_map.items())[:5]):
                print(f"  {invoice}: {date}")
            
            # Display info about Sheet2 invoice numbers
            sheet2_invoices = [str(inv).strip() if pd.notna(inv) else "None" for inv in sheet2_df['Invoice Number']]
            print(f"Found {len(sheet2_invoices)} invoice numbers in Sheet2")
            print("First 5 invoice numbers in Sheet2:")
            for i, inv in enumerate(sheet2_invoices[:5]):
                print(f"  {inv}")
            
            # Update Sheet2 Invoice Dates based on matching Invoice Numbers
            updated_count = 0
            for i, row in sheet2_df.iterrows():
                if pd.isna(row['Invoice Number']):
                    continue
                    
                # Process invoice number to match format in map
                sheet2_invoice = row['Invoice Number']
                
                # Convert to string if it's a number
                if isinstance(sheet2_invoice, (int, float)):
                    sheet2_invoice = str(int(sheet2_invoice))
                else:
                    sheet2_invoice = str(sheet2_invoice).strip()
                    
                if not sheet2_invoice:  # Skip if empty after processing
                    continue
                    
                # Check if invoice number exists in our mapping
                if sheet2_invoice in invoice_date_map:
                    # Get the date value from our map
                    date_value = invoice_date_map[sheet2_invoice]
                    
                    # Format the date as DD-MM-YYYY (remove any time component)
                    try:
                        # Parse the date with pandas (handles various formats)
                        parsed_date = pd.to_datetime(date_value)
                        # Format strictly as DD-MM-YYYY
                        formatted_date = parsed_date.strftime('%d-%m-%Y')
                        sheet2_df.at[i, 'Invoice Date'] = formatted_date
                        print(f"Formatted date: {formatted_date}")
                    except Exception as e:
                        print(f"Warning: Could not format date {date_value}: {e}")
                        # If parsing fails, use as is
                        sheet2_df.at[i, 'Invoice Date'] = date_value
                    
                    updated_count += 1
                    print(f"Updated invoice {sheet2_invoice} with date {sheet2_df.at[i, 'Invoice Date']}")
                    
            if updated_count == 0:
                messagebox.showinfo("Info", "No matching Invoice Numbers found to update Invoice Dates.")
                return 0
            
            # Save the updated file with proper formatting
            try:
                # First save the DataFrame to Excel
                with pd.ExcelWriter(consolidated_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    sheet2_df.to_excel(writer, sheet_name='Sheet2', index=False)
                
                # Then open the workbook to apply text formatting to date cells
                wb = load_workbook(consolidated_file)
                ws = wb['Sheet2']
                
                # Find the Invoice Date column
                date_col_idx = None
                for idx, col in enumerate(ws[1], 1):  # 1-based indexing for Excel columns
                    if col.value == 'Invoice Date':
                        date_col_idx = idx
                        break
                
                if date_col_idx:
                    # Apply text format to all cells in the Invoice Date column
                    for row in range(2, ws.max_row + 1):  # Skip header row
                        cell = ws.cell(row=row, column=date_col_idx)
                        if cell.value:
                            # Set as text to preserve the exact format
                            cell.number_format = '@'  # Text format
                
                # Save the workbook with formatting
                wb.save(consolidated_file)
                print("Applied text formatting to date cells")
            except Exception as e:
                print(f"Warning: Could not apply text formatting to Excel file: {e}")
                # The basic save was already done, so we continue
            
            print(f"\nSuccessfully updated {updated_count} Invoice Dates in Sheet2")
            print("Dates have been formatted as DD-MM-YYYY")
            messagebox.showinfo("Success", f"Successfully updated {updated_count} Invoice Dates in Sheet2")
            return updated_count
        
        except Exception as e:
            print(f"Error reading source file: {e}")
            messagebox.showerror("Error", f"Could not read source file: {e}")
            return 0
        
    except Exception as e:
        print(f"Error filling invoice dates: {e}")
        messagebox.showerror("Error", f"Failed to fill Invoice Dates: {e}")
        return 0

def main():
    try:
        # Ask user to select folder
        print("Please select the folder containing CSV files...")
        folder_path = select_folder()
        
        if not folder_path:
            print("No folder selected. Exiting...")
            return
        
        print(f"Selected folder: {folder_path}")
        print("Searching for CSV files in all subfolders...")
        
        # Find all combined_* CSV files
        csv_files = find_combined_files(folder_path)
        
        if not csv_files:
            print("No CSV files found starting with 'combined_' in any subfolder")
            return
        
        print(f"\nFound {len(csv_files)} CSV files to process:")
        for file in csv_files:
            print(f"- {file}")
        
        # Ask user preference
        append_to_existing = ask_user_preference()
        existing_df = None
        output_file = None
        
        if append_to_existing:
            print("\nPlease select the existing consolidated file to append to...")
            existing_file = select_existing_file()
            if not existing_file:
                print("No file selected for appending. Exiting...")
                return
            
            try:
                # Check if it's an Excel file
                if existing_file.endswith('.xlsx'):
                    existing_df = pd.read_excel(existing_file, sheet_name='Sheet1')
                    sheet2_df = pd.read_excel(existing_file, sheet_name='Sheet2')
                else:
                    existing_df = pd.read_csv(existing_file)
                    sheet2_df = create_sheet2_template()
                output_file = os.path.splitext(existing_file)[0] + '.xlsx'
                print(f"Successfully loaded existing file: {existing_file}")
            except Exception as e:
                print(f"Error reading existing file: {e}")
                return
        else:
            # Create consolidated folder and new output file path
            consolidated_folder = create_consolidated_folder(folder_path)
            timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(consolidated_folder, f"consolidated_data_{timestamp}.xlsx")
            sheet2_df = create_sheet2_template()
        
        # Combine CSV files
        combined_data = combine_csv_files(csv_files, existing_df)
        
        # Save combined data as Excel with two sheets
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            combined_data.to_excel(writer, sheet_name='Sheet1', index=False)
            sheet2_df.to_excel(writer, sheet_name='Sheet2', index=False)
        
        print(f"\nSuccessfully processed files!")
        print(f"Output saved to: {output_file}")
        print(f"Total rows in Sheet1: {len(combined_data)}")
        print(f"Sheet2 has been created with the required columns for manual data entry")
        
        # Print summary of data from each folder
        print("\nSummary of files processed by folder:")
        print("-" * 70)
        folder_summary = combined_data.groupby('Source_Folder').agg({
            'Source_File': 'nunique',
        }).reset_index()
        for _, row in folder_summary.iterrows():
            print(f"Folder: {row['Source_Folder']}")
            print(f"Number of files: {row['Source_File']}")
            print("-" * 70)
        
        # Fill Invoice Dates
        updated_count = fill_invoice_dates(output_file)
        if updated_count > 0:
            print("Invoice Dates filled successfully!")
        
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    main() 